import re
import fitz
import pdfplumber
from lxml import etree
import zipfile
import markdown
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from .utils import is_likely_heading, infer_heading_level, is_table_header

def extract_outline_from_markdown(md_path):
    """
    解析 Markdown 文件，按原顺序提取大纲信息和内容，与docx解析器格式对齐
    """
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            md_content = f.read()

        # 将 Markdown 转换为 HTML
        html_content = markdown.markdown(md_content, extensions=['extra'])

        # 使用 BeautifulSoup 解析 HTML
        soup = BeautifulSoup(html_content, 'html.parser')

        extracted_content = "Outline and Text:\n"

        # 获取所有需要处理的元素，保持原文档顺序
        elements = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p', 'ul', 'ol', 'blockquote', 'pre', 'table'])

        # 按原文档顺序处理每个元素
        for element in elements:
            if element.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                # 处理标题 - 与docx解析器保持一致的格式
                level = int(element.name[1])  # 获取标题级别
                title = element.get_text().strip()
                indent = "  " * (level - 1)
                extracted_content += f"{indent}Heading {level}: {title}\n"
            
            elif element.name == 'p':
                # 处理段落
                text = element.get_text().strip()
                if text:
                    extracted_content += f"Paragraph: {text}\n"
            
            elif element.name in ['ul', 'ol']:
                # 处理列表 - 与docx解析器保持一致，不添加特殊标记
                for li in element.find_all('li'):
                    text = li.get_text().strip()
                    if text:
                        extracted_content += f"Paragraph: {text}\n"
            
            elif element.name == 'blockquote':
                # 处理块引用
                text = element.get_text().strip()
                if text:
                    extracted_content += f"Paragraph: {text}\n"
            
            elif element.name == 'pre':
                # 处理代码块
                code = element.find('code')
                if code:
                    text = code.get_text().strip()
                    extracted_content += f"Paragraph: {text}\n"
            
            elif element.name == 'table':
                # 处理表格 - 与docx解析器格式对齐
                extracted_content += f"Table:\n"
                rows = element.find_all('tr')
                for row in rows:
                    cells = row.find_all(['th', 'td'])
                    row_data = []
                    for cell in cells:
                        cell_text = cell.get_text().strip()
                        row_data.append(cell_text)
                    if row_data:
                        extracted_content += "Paragraph: " + " | ".join(row_data) + "\n"

        return extracted_content

    except Exception as e:
        raise Exception(f"Error parsing Markdown file: {e}")
    

from openpyxl import load_workbook

def extract_outline_from_excel(xlsx_path):
    """
    解析 Excel 文件，提取工作表和單元格內容
    """
    try:
        workbook = load_workbook(xlsx_path)
        extracted_content = "Outline and Text:\n"

        # 處理每個工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 添加工作表作為一級標題
            extracted_content += f"Heading 1: sheet - {sheet_name}\n"

            # 提取單元格內容
            for row in sheet.iter_rows(values_only=True):
                row_data = []
                for cell in row:
                    if cell is not None:
                        row_data.append(str(cell))

                if row_data:
                    # 檢查是否可能是表頭
                    if is_table_header(row_data):
                        extracted_content += f"Heading 2: line\n"

                    # 添加行內容
                    row_text = " | ".join(row_data)
                    extracted_content += f"Paragraph: {row_text}\n"

        return extracted_content

    except Exception as e:
        raise Exception(f"Error parsing Excel file: {e}")



try:
    from langchain.text_splitter import RecursiveCharacterTextSplitter
except ImportError:
    print("Please install langchain: pip install langchain")
    raise

def extract_outline_from_pdf(pdf_path):
    """
    解析 PDF 文件，提取大綱信息並格式化為與 Word 一致的格式
    """
    try:
        # 使用 PyMuPDF 提取大綱
        doc = fitz.open(pdf_path)
        outline = doc.get_toc()

        # 使用 pdfplumber 提取文本內容
        with pdfplumber.open(pdf_path) as pdf:
            extracted_content = "Outline and Text:\n"

            # 如果有大綱，建立頁面到標題的映射
            page_to_headings = {}
            if outline:
                for item in outline:
                    level, title, page_num = item
                    if page_num not in page_to_headings:
                        page_to_headings[page_num] = []
                    page_to_headings[page_num].append((level, title))

            # 逐頁處理內容
            for page_num, page in enumerate(pdf.pages):
                current_page = page_num + 1

                # 如果當前頁有標題，先輸出標題
                if current_page in page_to_headings:
                    for level, title in page_to_headings[current_page]:
                        indent = "  " * (level - 1)
                        extracted_content += f"{indent}Heading {level}: {title}\n"

                # 提取頁面文本
                text = page.extract_text()
                if text:
                    # 將文本按段落分割
                    paragraphs = text.split('\n\n')
                    for paragraph in paragraphs:
                        # 清理段落文本
                        cleaned_paragraph = ' '.join(paragraph.split()).strip()
                        if cleaned_paragraph and len(cleaned_paragraph) > 10:  # 過濾掉太短的內容
                            # 檢查是否可能是標題（簡單的啟發式判斷）
                            if is_likely_heading(cleaned_paragraph):
                                # 嘗試推斷標題級別
                                level = infer_heading_level(cleaned_paragraph)
                                indent = "  " * (level - 1)
                                extracted_content += f"{indent}Heading {level}: {cleaned_paragraph}\n"
                            else:
                                extracted_content += f"Paragraph: {cleaned_paragraph}\n"

                # 提取表格
                tables = page.extract_tables()
                if tables:
                    for table_idx, table in enumerate(tables):
                        extracted_content += f"Table {table_idx + 1} (Page {current_page}):\n"
                        for row in table:
                            if row and any(cell for cell in row if cell):  # 確保行不為空
                                row_text = " | ".join(str(cell) if cell else "" for cell in row)
                                extracted_content += f"Paragraph: {row_text}\n"

        doc.close()
        return extracted_content

    except Exception as e:
        raise Exception(f"Error parsing PDF file: {e}")


def extract_outline_from_docx_with_tables(docx_path):
    """
    增強版 docx 解析，包含表格處理
    """
    try:
        with zipfile.ZipFile(docx_path, 'r') as docx_zip:
            document_xml = docx_zip.read('word/document.xml')

            # 嘗試讀取樣式文件，但有些文檔可能沒有
            styles_xml = None
            try:
                styles_xml = docx_zip.read('word/styles.xml')
            except KeyError:
                print(f"Warning: styles.xml not found in {docx_path}")
                pass

            doc_tree = etree.fromstring(document_xml)
            styles_tree = etree.fromstring(styles_xml) if styles_xml else None

            namespaces = {
                'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
            }

            # 獲取樣式映射
            style_outline_map = {}
            if styles_tree is not None:
                styles = styles_tree.xpath('//w:style', namespaces=namespaces)
                for style in styles:
                    style_id = style.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}styleId')
                    outline_lvl = style.xpath('.//w:outlineLvl', namespaces=namespaces)
                    if outline_lvl:
                        level = int(outline_lvl[0].get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val'))
                        style_outline_map[style_id] = level + 1
                    # 額外檢查：如果樣式名包含Heading但沒有outlineLvl，推斷級別
                    elif style_id and style_id.startswith('Heading'):
                        try:
                            level = int(re.findall(r'\d+', style_id)[0])
                            style_outline_map[style_id] = level
                        except (ValueError, IndexError):
                            pass

            extracted_content = "Outline and Text:\n"

            # 獲取所有段落和表格，按文檔順序處理
            body_elements = doc_tree.xpath('//w:body/*', namespaces=namespaces)

            for element in body_elements:
                if element.tag.endswith('p'):  # 段落
                    # 處理段落
                    text_elements = element.xpath('.//w:t', namespaces=namespaces)
                    paragraph_text = ''.join([t.text or '' for t in text_elements]).strip()

                    if not paragraph_text:
                        continue

                    # 獲取段落級別 - 多種方式嘗試
                    outline_level = None

                    # 方式1: 檢查pStyle
                    pStyle = element.xpath('.//w:pStyle', namespaces=namespaces)
                    if pStyle:
                        style_id = pStyle[0].get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                        if style_id in style_outline_map:
                            outline_level = style_outline_map[style_id]
                        elif style_id and (style_id.startswith('Heading') or style_id.startswith('heading')):
                            try:
                                outline_level = int(re.findall(r'\d+', style_id)[0])
                            except (ValueError, IndexError):
                                pass

                    # 方式2: 直接檢查outlineLvl
                    if outline_level is None:
                        outline_lvl_elem = element.xpath('.//w:outlineLvl', namespaces=namespaces)
                        if outline_lvl_elem:
                            outline_level = int(outline_lvl_elem[0].get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')) + 1

                    # 方式3: 啟發式判斷（如果前面幾種方式都失敗）
                    if outline_level is None and is_likely_heading(paragraph_text):
                        outline_level = infer_heading_level(paragraph_text)

                    # 格式化輸出
                    if outline_level:
                        indent = "  " * (outline_level - 1)
                        extracted_content += f"{indent}Heading {outline_level}: {paragraph_text}\n"
                    else:
                        extracted_content += f"Paragraph: {paragraph_text}\n"

                elif element.tag.endswith('tbl'):  # 表格
                    rows = element.xpath('.//w:tr', namespaces=namespaces)
                    if rows:
                        extracted_content += f"Table:\n"
                        for row_idx, row in enumerate(rows):
                            cells = row.xpath('.//w:tc', namespaces=namespaces)
                            row_data = []
                            for cell in cells:
                                cell_text = ''.join([t.text or '' for t in cell.xpath('.//w:t', namespaces=namespaces)])
                                row_data.append(cell_text.strip())

                            if row_data:
                                extracted_content += "Paragraph: " + " | ".join(row_data) + "\n"

            return extracted_content

    except Exception as e:
        raise Exception(f"Error parsing docx file: {e}")